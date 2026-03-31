import openpyxl
import json
import sys

print("Loading workbook...")
try:
    wb = openpyxl.load_workbook(r'd:\RAYSAN\KMTI Data Management\Systems\kmti_workstation_v3_new\backend\data\KMTI Raw Material Calculator.xlsx', data_only=True)
    print("Sheets:", wb.sheetnames)
    
    sheet = wb.active
    rows = []
    
    # Just read the first 10 rows to see the structure
    for count, row in enumerate(sheet.iter_rows(values_only=True)):
        if count >= 10:
            break
        rows.append(list(row))
        
    print("Top 10 rows:")
    for r in rows:
        print(r)
        
except Exception as e:
    print("Error:", e)
