import os
import re
import datetime
import unicodedata
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from io import BytesIO

from db.database import get_db, get_fms_db
from core.auth import get_current_user
from models.user import User, UserRole
from models.fms import FmsUser
from models.work_schedule import WorkScheduleJob, WorkScheduleComponent, WorkScheduleAssignment
from socket_manager import sio

router = APIRouter()

EXCEL_FILE_PATH = r"d:\RAYSAN\KMTI Data Management\Systems\kmti_workstation_v3_new\backend\data\KMTI Work Schedule Monitoring 2026.06.19.xlsx"

# --- Request/Response Pydantic Models ---

class ComponentUpdatePayload(BaseModel):
    status: str
    submitted_date: Optional[str] = None # format YYYY-MM-DD or None

class JobCreatePayload(BaseModel):
    job_id: str
    deadline: Optional[str] = None

class ComponentCreatePayload(BaseModel):
    unit_code: str
    assembly_3d: Optional[str] = "-"
    parts_3d: Optional[str] = "-"
    assembly_2d: Optional[str] = "-"
    parts_2d: Optional[str] = "-"
    status: Optional[str] = "Pending/Not Started"
    submitted_date: Optional[str] = None


class ComponentUpdateAllPayload(BaseModel):
    unit_code: Optional[str] = None
    assembly_3d: Optional[str] = None
    parts_3d: Optional[str] = None
    assembly_2d: Optional[str] = None
    parts_2d: Optional[str] = None
    status: Optional[str] = None
    submitted_date: Optional[str] = None


# --- Permission Dependency ---

async def require_schedule_write(
    current_user = Depends(get_current_user),
    fms_db: AsyncSession = Depends(get_fms_db)
):
    role_str = current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    if role_str in ("admin", "it"):
        return current_user
    
    # Check FMS DB role
    try:
        fms_res = await fms_db.execute(select(FmsUser).where(FmsUser.username == current_user.username))
        fms_user = fms_res.scalar_one_or_none()
        if fms_user and fms_user.role.upper() in ("TEAM LEADER", "LEADER", "ADMIN"):
            return current_user
    except Exception as e:
        import logging
        logger = logging.getLogger("uvicorn.error")
        logger.warning(f"Error validating Team Leader role in FMS DB: {e}")

    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Access denied. Only Admins and Team Leaders can edit the schedule."
    )

# --- Helper functions ---

def normalize_status(status_val):
    if not status_val:
        return "Pending/Not Started"
    s = str(status_val).strip().lower()
    if s in ("completed", "complete", "complete "):
        return "Completed"
    if "checking" in s:
        return "For Checking"
    if s == "-":
        return "Excluded/NA"
    return "Pending/Not Started"

# --- Routing Endpoints ---

@router.get("/jobs")
async def get_jobs(db: AsyncSession = Depends(get_db)):
    """
    Get all schedule jobs with progress metrics.
    """
    # Fetch all jobs
    jobs_res = await db.execute(select(WorkScheduleJob))
    jobs = jobs_res.scalars().all()
    
    # Fetch all components to calculate completion
    comp_res = await db.execute(select(WorkScheduleComponent))
    components = comp_res.scalars().all()
    
    # Group components by job_id
    comp_by_job = {}
    for c in components:
        if c.job_id not in comp_by_job:
            comp_by_job[c.job_id] = []
        comp_by_job[c.job_id].append(c)
        
    res = []
    for j in jobs:
        comps = comp_by_job.get(j.job_id, [])
        total = len(comps)
        completed = sum(1 for c in comps if normalize_status(c.status) == "Completed")
        checking = sum(1 for c in comps if normalize_status(c.status) == "For Checking")
        progress = (completed / total * 100) if total > 0 else 0.0
        
        res.append({
            "id": j.id,
            "job_id": j.job_id,
            "deadline": j.deadline,
            "total_components": total,
            "completed_components": completed,
            "checking_components": checking,
            "progress_percent": round(progress, 1)
        })
        
    # Sort by progress (ascending or descending) or alphabetically by job ID
    res.sort(key=lambda x: x["job_id"])
    return {"success": True, "jobs": res}


@router.get("/jobs/{job_id}/components")
async def get_components(job_id: str, db: AsyncSession = Depends(get_db)):
    """
    Get all components for a specific job.
    """
    comp_res = await db.execute(
        select(WorkScheduleComponent).where(WorkScheduleComponent.job_id == job_id)
    )
    components = comp_res.scalars().all()
    
    res = [
        {
            "id": c.id,
            "job_id": c.job_id,
            "unit_code": c.unit_code,
            "assembly_3d": c.assembly_3d,
            "parts_3d": c.parts_3d,
            "assembly_2d": c.assembly_2d,
            "parts_2d": c.parts_2d,
            "status": c.status,
            "submitted_date": c.submitted_date.strftime("%Y-%m-%d") if c.submitted_date else None
        }
        for c in components
    ]
    return {"success": True, "components": res}


@router.post("/components/{component_id}/status")
async def update_component_status(
    component_id: int,
    payload: ComponentUpdatePayload,
    db: AsyncSession = Depends(get_db),
    user = Depends(require_schedule_write)
):
    """
    Update status and submitted date of a component.
    """
    comp = await db.get(WorkScheduleComponent, component_id)
    if not comp:
        raise HTTPException(status_code=404, detail="Component not found.")
        
    comp.status = payload.status
    if payload.submitted_date:
        try:
            comp.submitted_date = datetime.datetime.strptime(payload.submitted_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Expected YYYY-MM-DD.")
    else:
        comp.submitted_date = None
        
    await db.commit()
    await sio.emit('schedule_updated')
    return {
        "success": True, 
        "message": f"Component '{comp.unit_code}' updated successfully.",
        "component": {
            "id": comp.id,
            "status": comp.status,
            "submitted_date": comp.submitted_date.strftime("%Y-%m-%d") if comp.submitted_date else None
        }
    }


@router.post("/jobs")
async def create_job(
    payload: JobCreatePayload,
    db: AsyncSession = Depends(get_db),
    user = Depends(require_schedule_write)
):
    """
    Create a new Job Group.
    """
    # Check if job exists
    res = await db.execute(select(WorkScheduleJob).where(WorkScheduleJob.job_id == payload.job_id))
    if res.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Job ID already exists.")
        
    new_job = WorkScheduleJob(
        job_id=payload.job_id,
        deadline=payload.deadline
    )
    db.add(new_job)
    await db.commit()
    await db.refresh(new_job)
    await sio.emit('schedule_updated')
    return {"success": True, "job": {"id": new_job.id, "job_id": new_job.job_id, "deadline": new_job.deadline}}


@router.post("/jobs/{job_id}/components")
async def create_component(
    job_id: str,
    payload: ComponentCreatePayload,
    db: AsyncSession = Depends(get_db),
    user = Depends(require_schedule_write)
):
    """
    Create a new component unit in a job.
    """
    job_res = await db.execute(select(WorkScheduleJob).where(WorkScheduleJob.job_id == job_id))
    job = job_res.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")
        
    sub_date = None
    if payload.submitted_date:
        try:
            sub_date = datetime.datetime.strptime(payload.submitted_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Expected YYYY-MM-DD.")
            
    new_comp = WorkScheduleComponent(
        job_id=job_id,
        unit_code=payload.unit_code,
        assembly_3d=payload.assembly_3d,
        parts_3d=payload.parts_3d,
        assembly_2d=payload.assembly_2d,
        parts_2d=payload.parts_2d,
        status=payload.status,
        submitted_date=sub_date
    )
    db.add(new_comp)
    await db.commit()
    await db.refresh(new_comp)
    await sio.emit('schedule_updated')
    return {"success": True, "component": {"id": new_comp.id, "unit_code": new_comp.unit_code}}


@router.delete("/jobs/{job_id}")
async def delete_job(
    job_id: str,
    db: AsyncSession = Depends(get_db),
    user = Depends(require_schedule_write)
):
    """
    Delete a Job Group and cascade delete all its components.
    """
    job_res = await db.execute(select(WorkScheduleJob).where(WorkScheduleJob.job_id == job_id))
    job = job_res.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")
        
    await db.delete(job)
    await db.commit()
    await sio.emit('schedule_updated')
    return {"success": True, "message": f"Job '{job_id}' deleted successfully."}


@router.delete("/components/{component_id}")
async def delete_component(
    component_id: int,
    db: AsyncSession = Depends(get_db),
    user = Depends(require_schedule_write)
):
    """
    Delete a specific drawing component unit.
    """
    comp = await db.get(WorkScheduleComponent, component_id)
    if not comp:
        raise HTTPException(status_code=404, detail="Component not found.")
        
    await db.delete(comp)
    await db.commit()
    await sio.emit('schedule_updated')
    return {"success": True, "message": "Component deleted successfully."}


@router.patch("/components/{component_id}")
async def patch_component(
    component_id: int,
    payload: ComponentUpdateAllPayload,
    db: AsyncSession = Depends(get_db),
    user = Depends(require_schedule_write)
):
    """
    Modify any details of a drawing component unit.
    """
    comp = await db.get(WorkScheduleComponent, component_id)
    if not comp:
        raise HTTPException(status_code=404, detail="Component not found.")
        
    if payload.unit_code is not None:
        comp.unit_code = payload.unit_code
    if payload.assembly_3d is not None:
        comp.assembly_3d = payload.assembly_3d
    if payload.parts_3d is not None:
        comp.parts_3d = payload.parts_3d
    if payload.assembly_2d is not None:
        comp.assembly_2d = payload.assembly_2d
    if payload.parts_2d is not None:
        comp.parts_2d = payload.parts_2d
    if payload.status is not None:
        comp.status = payload.status
    if payload.submitted_date is not None:
        if payload.submitted_date:
            try:
                comp.submitted_date = datetime.datetime.strptime(payload.submitted_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Expected YYYY-MM-DD.")
        else:
            comp.submitted_date = None
            
    await db.commit()
    await sio.emit('schedule_updated')
    return {
        "success": True, 
        "message": "Component updated successfully.",
        "component": {
            "id": comp.id,
            "unit_code": comp.unit_code,
            "status": comp.status,
            "submitted_date": comp.submitted_date.strftime("%Y-%m-%d") if comp.submitted_date else None
        }
    }


@router.post("/import")
async def import_from_excel(
    db: AsyncSession = Depends(get_db),
    user = Depends(require_schedule_write)
):
    """
    Seed/Import work schedule data from the local Excel spreadsheet.
    DANGER: Clears existing work schedule tables.
    """
    if not os.path.exists(EXCEL_FILE_PATH):
        raise HTTPException(status_code=404, detail="Source Excel file not found on server.")
        
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
        ws = wb['Schedule']
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open Excel workbook: {e}")
        
    jobs_dict = {}
    current_job = None
    
    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_a_str = str(val_a).strip() if val_a is not None else ""
        
        if "Job Status" in val_a_str:
            # Parse Job ID
            m = re.match(r"(\d+)\s+Job\s+Status", val_a_str, re.IGNORECASE)
            job_id = m.group(1) if m else val_a_str.replace("Job Status", "").strip()
            
            # Deadline in col B
            deadline_val = ws.cell(row=r, column=2).value
            deadline_str = str(deadline_val).strip() if deadline_val is not None else ""
            
            current_job = {
                "deadline": deadline_str,
                "components": []
            }
            jobs_dict[job_id] = current_job
            
        elif current_job:
            if val_a_str == "" or val_a_str == "Machine/Unit Code" or val_a_str == "Legend:":
                continue
                
            unit_code = val_a_str
            assembly_3d = str(ws.cell(row=r, column=2).value or "").strip()
            parts_3d = str(ws.cell(row=r, column=5).value or "").strip()
            assembly_2d = str(ws.cell(row=r, column=7).value or "").strip()
            parts_2d = str(ws.cell(row=r, column=10).value or "").strip()
            status_val = str(ws.cell(row=r, column=12).value or "").strip()
            submitted_val = ws.cell(row=r, column=16).value
            
            sub_date = None
            if submitted_val:
                if isinstance(submitted_val, (datetime.datetime, datetime.date)):
                    sub_date = submitted_val.date() if isinstance(submitted_val, datetime.datetime) else submitted_val
                else:
                    try:
                        # parse standard YYYY-MM-DD
                        sub_date = datetime.datetime.strptime(str(submitted_val).split()[0], "%Y-%m-%d").date()
                    except:
                        pass
                        
            current_job["components"].append({
                "unit_code": unit_code,
                "assembly_3d": assembly_3d if assembly_3d else "-",
                "parts_3d": parts_3d if parts_3d else "-",
                "assembly_2d": assembly_2d if assembly_2d else "-",
                "parts_2d": parts_2d if parts_2d else "-",
                "status": status_val if status_val else "Pending/Not Started",
                "submitted_date": sub_date
            })
            
    # Clear and insert into DB
    try:
        await db.execute(delete(WorkScheduleComponent))
        await db.execute(delete(WorkScheduleJob))
        await db.commit()
        
        for jid, job_info in jobs_dict.items():
            db_job = WorkScheduleJob(job_id=jid, deadline=job_info["deadline"])
            db.add(db_job)
            await db.commit()
            await db.refresh(db_job)
            
            for comp in job_info["components"]:
                db_comp = WorkScheduleComponent(
                    job_id=jid,
                    unit_code=comp["unit_code"],
                    assembly_3d=comp["assembly_3d"],
                    parts_3d=comp["parts_3d"],
                    assembly_2d=comp["assembly_2d"],
                    parts_2d=comp["parts_2d"],
                    status=comp["status"],
                    submitted_date=comp["submitted_date"]
                )
                db.add(db_comp)
            await db.commit()
            
        await sio.emit('schedule_updated')
        return {"success": True, "message": f"Successfully imported {len(jobs_dict)} jobs from spreadsheet."}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Database import transaction failed: {e}")


@router.get("/export")
async def export_to_excel(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Export the current database state back to the Excel file by updating the 
    master file in-place, preserving the original layout, formatting, formulas, and Gantt charts.
    Streams the Excel file to the client for download.
    """
    if not os.path.exists(EXCEL_FILE_PATH):
        raise HTTPException(status_code=404, detail="Source Excel file not found on server.")
        
    try:
        # Load without data_only=True to preserve formulas and formatting intact!
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=False)
        ws = wb['Schedule']
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load master Excel template: {e}")

    # Fetch all database components
    comp_res = await db.execute(select(WorkScheduleComponent))
    db_components = comp_res.scalars().all()
    
    # Fetch all database jobs to track which ones need appending
    jobs_res = await db.execute(select(WorkScheduleJob))
    db_jobs = jobs_res.scalars().all()
    
    # Index components by (job_id, unit_code) in lowercase for mapping
    db_map = {}
    for c in db_components:
        key = (str(c.job_id).strip().lower(), str(c.unit_code).strip().lower())
        db_map[key] = c

    # Scan the worksheet in-place and update values
    current_job_id = None
    seen_jobs = set()
    
    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_a_str = str(val_a).strip() if val_a is not None else ""
        
        # Check if row is a job header
        if "Job Status" in val_a_str:
            # Parse Job ID
            m = re.match(r"(\d+)\s+Job\s+Status", val_a_str, re.IGNORECASE)
            current_job_id = m.group(1) if m else val_a_str.replace("Job Status", "").strip()
            seen_jobs.add(current_job_id.strip().lower())
            
        elif current_job_id:
            # Check if this row represents a component
            if val_a_str == "" or val_a_str == "Machine/Unit Code" or val_a_str == "Legend:":
                continue
                
            unit_code = val_a_str
            key = (str(current_job_id).strip().lower(), str(unit_code).strip().lower())
            
            if key in db_map:
                db_comp = db_map[key]
                # Update Status (Column L, index 12)
                ws.cell(row=r, column=12, value=db_comp.status)
                # Update Submitted Date (Column P, index 16)
                if db_comp.submitted_date:
                    ws.cell(row=r, column=16, value=db_comp.submitted_date.strftime("%Y-%m-%d"))
                else:
                    ws.cell(row=r, column=16, value=None)

    # Find reference template rows dynamically from existing sheet contents
    ref_header_row = None
    ref_comp_row = None
    
    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_a_str = str(val_a).strip() if val_a is not None else ""
        if "Job Status" in val_a_str:
            ref_header_row = r
        status_val = ws.cell(row=r, column=12).value
        if status_val and str(status_val).strip().lower() in ("pending/not started", "complete", "for checking"):
            ref_comp_row = r
            
    if not ref_header_row:
        ref_header_row = 1212
    if not ref_comp_row:
        ref_comp_row = 1216

    # Helper function to copy styles cell-by-cell
    from copy import copy
    def copy_cell_style(src_cell, dst_cell):
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

    # Find the actual last populated row of the sheet (avoiding empty styled rows)
    last_used_row = 1
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, 20)):
            last_used_row = r
            break
            
    next_row = last_used_row + 1

    # Append any jobs/components added in database but not present in the Excel template
    for j in db_jobs:
        j_key = j.job_id.strip().lower()
        if j_key not in seen_jobs:
            # Append empty separator row
            next_row += 1
            
            # Append Job Header row (e.g. 6969 Job Status)
            for c in range(1, 20):
                copy_cell_style(ws.cell(row=ref_header_row, column=c), ws.cell(row=next_row, column=c))
            ws.cell(row=next_row, column=1, value=f"{j.job_id} Job Status")
            ws.cell(row=next_row, column=2, value=f"Deadline: {j.deadline}" if j.deadline else "Deadline:")
            if ws.row_dimensions[ref_header_row].height:
                ws.row_dimensions[next_row].height = ws.row_dimensions[ref_header_row].height
            next_row += 1
            
            # Append Subheader 1 (e.g. Machine/Unit Code, 3D, 2D, Status, Submitted Date)
            for c in range(1, 20):
                copy_cell_style(ws.cell(row=ref_header_row + 1, column=c), ws.cell(row=next_row, column=c))
            ws.cell(row=next_row, column=1, value="Machine/Unit Code")
            ws.cell(row=next_row, column=2, value="3D")
            ws.cell(row=next_row, column=7, value="2D")
            ws.cell(row=next_row, column=12, value="Status")
            ws.cell(row=next_row, column=16, value="Submitted Date")
            if ws.row_dimensions[ref_header_row + 1].height:
                ws.row_dimensions[next_row].height = ws.row_dimensions[ref_header_row + 1].height
            
            # Merge Subheader 1 cells
            ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row + 1, end_column=1)
            ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=6)
            ws.merge_cells(start_row=next_row, start_column=7, end_row=next_row, end_column=11)
            ws.merge_cells(start_row=next_row, start_column=12, end_row=next_row + 1, end_column=15)
            ws.merge_cells(start_row=next_row, start_column=16, end_row=next_row, end_column=19)
            next_row += 1
            
            # Append Subheader 2 (e.g. Assembly, Parts, Assembly, Parts)
            for c in range(1, 20):
                copy_cell_style(ws.cell(row=ref_header_row + 2, column=c), ws.cell(row=next_row, column=c))
            ws.cell(row=next_row, column=2, value="Assembly")
            ws.cell(row=next_row, column=5, value="Parts")
            ws.cell(row=next_row, column=7, value="Assembly")
            ws.cell(row=next_row, column=10, value="Parts")
            ws.cell(row=next_row, column=16, value="Date")
            if ws.row_dimensions[ref_header_row + 2].height:
                ws.row_dimensions[next_row].height = ws.row_dimensions[ref_header_row + 2].height
                
            # Merge Subheader 2 cells
            ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=4)
            ws.merge_cells(start_row=next_row, start_column=5, end_row=next_row, end_column=6)
            ws.merge_cells(start_row=next_row, start_column=7, end_row=next_row, end_column=9)
            ws.merge_cells(start_row=next_row, start_column=10, end_row=next_row, end_column=11)
            ws.merge_cells(start_row=next_row, start_column=16, end_row=next_row, end_column=19)
            next_row += 1
            
            # Append component rows for this job
            j_comps = [c for c in db_components if str(c.job_id).strip().lower() == j_key]
            for c_item in j_comps:
                for c in range(1, 20):
                    copy_cell_style(ws.cell(row=ref_comp_row, column=c), ws.cell(row=next_row, column=c))
                
                ws.cell(row=next_row, column=1, value=c_item.unit_code)
                ws.cell(row=next_row, column=2, value=c_item.assembly_3d)
                ws.cell(row=next_row, column=5, value=c_item.parts_3d)
                ws.cell(row=next_row, column=7, value=c_item.assembly_2d)
                ws.cell(row=next_row, column=10, value=c_item.parts_2d)
                ws.cell(row=next_row, column=12, value=c_item.status)
                if c_item.submitted_date:
                    ws.cell(row=next_row, column=16, value=c_item.submitted_date.strftime("%Y-%m-%d"))
                else:
                    ws.cell(row=next_row, column=16, value=None)
                    
                if ws.row_dimensions[ref_comp_row].height:
                    ws.row_dimensions[next_row].height = ws.row_dimensions[ref_comp_row].height
                    
                # Merge component cells
                ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=4)
                ws.merge_cells(start_row=next_row, start_column=5, end_row=next_row, end_column=6)
                ws.merge_cells(start_row=next_row, start_column=7, end_row=next_row, end_column=9)
                ws.merge_cells(start_row=next_row, start_column=10, end_row=next_row, end_column=11)
                ws.merge_cells(start_row=next_row, start_column=12, end_row=next_row, end_column=15)
                ws.merge_cells(start_row=next_row, start_column=16, end_row=next_row, end_column=19)
                next_row += 1


    # 2. Fetch and write all database Gantt timeline assignments (System-first)
    assign_res = await db.execute(select(WorkScheduleAssignment))
    db_assignments = assign_res.scalars().all()
    
    # Map member name -> row number in the spreadsheet (Rows 5 to 10)
    member_rows = {}
    for r in range(5, 11):
        name_val = ws.cell(row=r, column=8).value
        if name_val:
            norm_name = unicodedata.normalize('NFC', str(name_val).strip().lower())
            member_rows[norm_name] = r
            
    # Write assignments to the cells
    for a in db_assignments:
        member_key = unicodedata.normalize('NFC', a.member_name.strip().lower())
        if member_key in member_rows:
            target_row = member_rows[member_key]
            ws.cell(row=target_row, column=a.col_index, value=a.value if a.value else None)
        
    # Overwrite the server Excel file to keep it synced (fail-safe if locked/open)
    try:
        wb.save(EXCEL_FILE_PATH)
    except Exception as e:
        import logging
        logger = logging.getLogger("uvicorn.error")
        logger.warning(f"Could not overwrite source file {EXCEL_FILE_PATH} (spreadsheet may be open by user): {e}")
        
    # Stream back as response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="KMTI Work Schedule Monitoring 2026.06.19.xlsx"'
    }
    return StreamingResponse(
        output,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



class TimelineUpdatePayload(BaseModel):
    member_name: str
    col_index: int
    value: str


@router.get("/timeline")
async def get_timeline(db: AsyncSession = Depends(get_db)):
    """
    Parses the Gantt chart timeline from rows 5-10, columns 20 onwards of the Excel sheet.
    """
    if not os.path.exists(EXCEL_FILE_PATH):
        raise HTTPException(status_code=404, detail="Source Excel file not found on server.")
        
    try:
        # Load with data_only=True to get final computed string values
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
        ws = wb['Schedule']
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load spreadsheet: {e}")
        
    # Get members (Rows 5 to 10)
    members = []
    for r in range(5, 11):
        name_val = ws.cell(row=r, column=8).value
        if name_val:
            members.append({
                "row": r,
                "name": str(name_val).strip()
            })
            
    # Find the last column that has a day number in row 4 (day headers)
    max_gantt_col = 20
    for c in range(ws.max_column, 19, -1):
        if ws.cell(row=4, column=c).value is not None:
            max_gantt_col = c
            break
            
    # Load database assignments
    db_res = await db.execute(select(WorkScheduleAssignment))
    db_assignments = db_res.scalars().all()
    db_map = {}
    for a in db_assignments:
        db_map[(a.member_name.strip().lower(), a.col_index)] = a.value
        
    timeline_days = []
    
    # Capture columns from 20 to max_gantt_col
    for c in range(20, max_gantt_col + 1):
        day_num = ws.cell(row=4, column=c).value
        day_week = ws.cell(row=3, column=c).value
        
        # Backtrack to find month
        month_name = ""
        for col_idx in range(c, 19, -1):
            m_val = ws.cell(row=2, column=col_idx).value
            if m_val:
                month_name = str(m_val).strip()
                break
                
        # Get assignments for each member on this column
        assignments = {}
        for m in members:
            key = (m["name"].strip().lower(), c)
            if key in db_map:
                assignments[m["name"]] = db_map[key] if db_map[key] is not None else ""
            else:
                cell_val = ws.cell(row=m["row"], column=c).value
                assignments[m["name"]] = str(cell_val).strip() if cell_val is not None else ""
            
        timeline_days.append({
            "col_index": c,
            "month": month_name,
            "day": day_num,
            "weekday": day_week,
            "assignments": assignments
        })
        
    return {
        "success": True,
        "members": [m["name"] for m in members],
        "timeline": timeline_days
    }


@router.post("/timeline")
async def update_timeline_cell(
    payload: TimelineUpdatePayload,
    user = Depends(require_schedule_write),
    db: AsyncSession = Depends(get_db)
):
    """
    Update a single Gantt chart cell in the database.
    """
    if not os.path.exists(EXCEL_FILE_PATH):
        raise HTTPException(status_code=404, detail="Source Excel file not found on server.")
        
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)
        ws = wb['Schedule']
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load spreadsheet: {e}")
        
    # Find the row for this member
    target_row = None
    for r in range(5, 11):
        name_val = ws.cell(row=r, column=8).value
        if name_val:
            normalized_name_val = unicodedata.normalize('NFC', str(name_val).strip().lower())
            normalized_payload_name = unicodedata.normalize('NFC', payload.member_name.strip().lower())
            if normalized_name_val == normalized_payload_name:
                target_row = r
                break
            
    if not target_row:
        raise HTTPException(status_code=404, detail=f"Member '{payload.member_name}' not found in Excel legend.")
        
    # Find existing assignment in database or create new one
    stmt = select(WorkScheduleAssignment).where(
        WorkScheduleAssignment.member_name == payload.member_name,
        WorkScheduleAssignment.col_index == payload.col_index
    )
    res = await db.execute(stmt)
    assignment = res.scalar_one_or_none()
    
    if assignment:
        assignment.value = payload.value if payload.value else None
    else:
        assignment = WorkScheduleAssignment(
            member_name=payload.member_name,
            col_index=payload.col_index,
            value=payload.value if payload.value else None
        )
        db.add(assignment)
        
    await db.commit()
    await sio.emit('schedule_updated')
    return {"success": True, "message": "Timeline cell updated in database."}


class TimelineSpanPayload(BaseModel):
    member_name: str
    start_col: int
    end_col: int
    job_code: str


@router.post("/timeline/span")
async def update_timeline_span(
    payload: TimelineSpanPayload,
    user = Depends(require_schedule_write),
    db: AsyncSession = Depends(get_db)
):
    """
    Update a range of Gantt chart cells representing a duration span in the database.
    """
    if not os.path.exists(EXCEL_FILE_PATH):
        raise HTTPException(status_code=404, detail="Source Excel file not found on server.")
        
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)
        ws = wb['Schedule']
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load spreadsheet: {e}")
        
    # Find the row for this member
    target_row = None
    for r in range(5, 11):
        name_val = ws.cell(row=r, column=8).value
        if name_val:
            normalized_name_val = unicodedata.normalize('NFC', str(name_val).strip().lower())
            normalized_payload_name = unicodedata.normalize('NFC', payload.member_name.strip().lower())
            if normalized_name_val == normalized_payload_name:
                target_row = r
                break
            
    if not target_row:
        raise HTTPException(status_code=404, detail=f"Member '{payload.member_name}' not found in Excel legend.")
        
    start = min(payload.start_col, payload.end_col)
    end = max(payload.start_col, payload.end_col)
    
    # Query existing database assignments in the range
    stmt = select(WorkScheduleAssignment).where(
        WorkScheduleAssignment.member_name == payload.member_name,
        WorkScheduleAssignment.col_index >= start,
        WorkScheduleAssignment.col_index <= end
    )
    res = await db.execute(stmt)
    existing_assignments = {a.col_index: a for a in res.scalars().all()}
    
    center = (start + end) // 2
    for c in range(start, end + 1):
        if not payload.job_code.strip():
            val = None
        else:
            if c == center:
                val = payload.job_code
            else:
                val = "-->"
            
        if c in existing_assignments:
            existing_assignments[c].value = val
        else:
            new_assign = WorkScheduleAssignment(
                member_name=payload.member_name,
                col_index=c,
                value=val
            )
            db.add(new_assign)
            
    await db.commit()
    await sio.emit('schedule_updated')
    return {"success": True, "message": "Timeline span updated in database."}


