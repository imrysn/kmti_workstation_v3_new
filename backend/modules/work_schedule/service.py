import os
import re
import datetime
import unicodedata
import asyncio
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO
import openpyxl
import openpyxl.utils
from copy import copy

ET.register_namespace('xdr', 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing')
ET.register_namespace('a', 'http://schemas.openxmlformats.org/drawingml/2006/main')
ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')

def update_drawing_xml(drawing_bytes: bytes, assignments: List[Any], member_rows_map: Dict[str, int], col_widths_map: Optional[Dict[int, float]] = None) -> bytes:
    root = ET.fromstring(drawing_bytes)
    if col_widths_map is None:
        col_widths_map = {}
    
    # Keep anchors that do not belong to member timeline rows (rows 5 to 25, i.e. 0-index 4 to 24)
    anchors_to_keep = []
    for anchor in root.findall('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor'):
        from_elem = anchor.find('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from')
        if from_elem is not None:
            r_elem = from_elem.find('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row')
            if r_elem is not None:
                r_idx = int(r_elem.text)
                if r_idx in range(4, 60):
                    continue
        anchors_to_keep.append(anchor)
        
    root.clear()
    for a_elem in anchors_to_keep:
        root.append(a_elem)
        
    # Group assignments by member
    by_member = {}
    for a in assignments:
        if a.member_name == '__day_status__' or not a.value:
            continue
        m_norm = unicodedata.normalize('NFC', a.member_name.strip().lower())
        if m_norm in member_rows_map:
            by_member.setdefault(m_norm, []).append(a)
            
    shape_id = 5000
    for m_norm, ass_list in by_member.items():
        row_0 = member_rows_map[m_norm] - 1
        ass_list.sort(key=lambda x: x.col_index)
        
        # Group into contiguous blocks
        blocks = []
        if ass_list:
            curr = [ass_list[0]]
            for next_a in ass_list[1:]:
                if next_a.col_index - curr[-1].col_index <= 1:
                    curr.append(next_a)
                else:
                    blocks.append(curr)
                    curr = [next_a]
            blocks.append(curr)
            
        for block in blocks:
            job_a = next((x for x in block if x.value not in ('-->', '->', '---')), None)
            if not job_a:
                continue
            
            start_col_0 = block[0].col_index - 1
            end_col_0 = block[-1].col_index - 1
            job_label = str(job_a.value).strip()
            
            end_col_num = block[-1].col_index
            end_col_w = col_widths_map.get(end_col_num, 3.50)
            end_col_off = int(end_col_w * 65000)
            
            shape_id += 1
            # 1. Straight Arrow Connector Shape (lowered to rowOff 195000 near bottom of row)
            arrow_xml = f'''<xdr:twoCellAnchor xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:from><xdr:col>{start_col_0}</xdr:col><xdr:colOff>30000</xdr:colOff><xdr:row>{row_0}</xdr:row><xdr:rowOff>195000</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>{end_col_0}</xdr:col><xdr:colOff>{end_col_off}</xdr:colOff><xdr:row>{row_0}</xdr:row><xdr:rowOff>195000</xdr:rowOff></xdr:to>
  <xdr:cxnSp macro="">
    <xdr:nvCxnSpPr>
      <xdr:cNvPr id="{shape_id}" name="Straight Arrow Connector {shape_id}"/>
      <xdr:cNvCxnSpPr/>
    </xdr:nvCxnSpPr>
    <xdr:spPr>
      <a:prstGeom prst="straightConnector1"><a:avLst/></a:prstGeom>
      <a:ln w="19050">
        <a:solidFill><a:srgbClr val="0070C0"/></a:solidFill>
        <a:prstDash val="dash"/>
        <a:tailEnd type="arrow" w="med" len="med"/>
      </a:ln>
    </xdr:spPr>
    <xdr:style>
      <a:lnRef idx="1"><a:schemeClr val="accent1"/></a:lnRef>
      <a:fillRef idx="0"><a:schemeClr val="accent1"/></a:fillRef>
      <a:effectRef idx="0"><a:schemeClr val="accent1"/></a:effectRef>
      <a:fontRef idx="minor"><a:schemeClr val="tx1"/></a:fontRef>
    </xdr:style>
  </xdr:cxnSp>
  <xdr:clientData/>
</xdr:twoCellAnchor>'''
            root.append(ET.fromstring(arrow_xml))
            
            shape_id += 1
            # 2. Text Box Shape (rowOff 30000 to 140000 - lowered from top border, 55000 EMU gap to arrow line)
            tb_xml = f'''<xdr:twoCellAnchor xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:from><xdr:col>{start_col_0}</xdr:col><xdr:colOff>10000</xdr:colOff><xdr:row>{row_0}</xdr:row><xdr:rowOff>30000</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>{end_col_0}</xdr:col><xdr:colOff>{end_col_off}</xdr:colOff><xdr:row>{row_0}</xdr:row><xdr:rowOff>140000</xdr:rowOff></xdr:to>
  <xdr:sp macro="" textlink="">
    <xdr:nvSpPr>
      <xdr:cNvPr id="{shape_id}" name="TextBox {shape_id}"/>
      <xdr:cNvSpPr txBox="1"/>
    </xdr:nvSpPr>
    <xdr:spPr>
      <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
      <a:noFill/>
      <a:ln w="9525" cmpd="sng"><a:noFill/></a:ln>
    </xdr:spPr>
    <xdr:style>
      <a:lnRef idx="0"><a:scrgbClr r="0" g="0" b="0"/></a:lnRef>
      <a:fillRef idx="0"><a:scrgbClr r="0" g="0" b="0"/></a:fillRef>
      <a:effectRef idx="0"><a:scrgbClr r="0" g="0" b="0"/></a:effectRef>
      <a:fontRef idx="minor"><a:schemeClr val="dk1"/></a:fontRef>
    </xdr:style>
    <xdr:txBody>
      <a:bodyPr vertOverflow="clip" wrap="none" lIns="0" tIns="0" rIns="0" bIns="0" anchor="t"/>
      <a:lstStyle/>
      <a:p>
        <a:pPr algn="ctr"/>
        <a:r>
          <a:rPr kumimoji="1" lang="en-US" altLang="ja-JP" sz="850" b="1">
            <a:solidFill><a:srgbClr val="FF0000"/></a:solidFill>
          </a:rPr>
          <a:t>{job_label}</a:t>
        </a:r>
      </a:p>
    </xdr:txBody>
  </xdr:sp>
  <xdr:clientData/>
</xdr:twoCellAnchor>'''
            root.append(ET.fromstring(tb_xml))
            
    return ET.tostring(root, encoding='utf-8', xml_declaration=True)

def determine_year(month_str, day_num, weekday_str) -> int:
    try:
        day_val = int(day_num)
    except (ValueError, TypeError):
        return 2026

    wd_map = {
        'mon': 0, 'tue': 1, 'wed': 2, 'thu': 3, 'fri': 4, 'sat': 5, 'sun': 6
    }
    wd_target = wd_map.get(str(weekday_str).strip().lower()[:3])
    if wd_target is None:
        return 2026

    month_map = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    m_idx = month_map.get(str(month_str).strip().lower())
    if not m_idx:
        return 2026

    current_year = datetime.date.today().year
    for y in range(current_year - 3, current_year + 4):
        try:
            dt = datetime.date(y, m_idx, day_val)
            if dt.weekday() == wd_target:
                return y
        except ValueError:
            continue
    return 2026

class ExcelScheduleService:
    # Static cache to store the layout structure of the timeline (members and columns)
    # This avoids reloading the heavy Excel workbook on every timeline read.
    _cached_layout: Optional[Dict[str, Any]] = None
    _cached_template_bytes: Optional[bytes] = None   # raw bytes of the template file
    _lock = asyncio.Lock()

    @classmethod
    def clear_cache(cls):
        cls._cached_layout = None
        cls._cached_template_bytes = None  # force re-read on next import/export

    @classmethod
    async def get_target_row_for_member(cls, excel_path: str, member_name: str) -> Optional[int]:
        if cls._cached_layout is None:
            async with cls._lock:
                if cls._cached_layout is None:
                    cls._cached_layout = await asyncio.to_thread(cls._load_and_parse_layout, excel_path)
        
        norm_member_name = unicodedata.normalize('NFC', member_name.strip().lower())
        return cls._cached_layout.get("member_rows", {}).get(norm_member_name)

    @staticmethod
    def parse_excel_for_import(excel_path: str) -> Dict[str, Any]:
        # Import clears and resets the DB, so we invalidate the layout cache
        ExcelScheduleService.clear_cache()
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Schedule']
        
        jobs_dict = {}
        current_job = None
        
        for r in range(1, ws.max_row + 1):
            val_a = ws.cell(row=r, column=1).value
            val_a_str = str(val_a).strip() if val_a is not None else ""
            
            if "Job Status" in val_a_str:
                m = re.match(r"(\d+)\s+Job\s+Status", val_a_str, re.IGNORECASE)
                job_id = m.group(1) if m else val_a_str.replace("Job Status", "").strip()
                
                deadline_val = ws.cell(row=r, column=2).value
                deadline_str = str(deadline_val).strip() if deadline_val is not None else ""
                
                current_job = {
                    "deadline": deadline_str,
                    "components": []
                }
                jobs_dict[job_id] = current_job
                
            elif current_job:
                if val_a_str == "" or val_a_str == "Machine/Unit Code" or val_a_str == "Legend:":
                    continue
                    
                unit_code = val_a_str
                assembly_3d = str(ws.cell(row=r, column=2).value or "").strip()
                parts_3d = str(ws.cell(row=r, column=5).value or "").strip()
                assembly_2d = str(ws.cell(row=r, column=7).value or "").strip()
                parts_2d = str(ws.cell(row=r, column=10).value or "").strip()
                status_val = str(ws.cell(row=r, column=12).value or "").strip()
                submitted_val = ws.cell(row=r, column=16).value
                
                sub_date = None
                if submitted_val:
                    if isinstance(submitted_val, (datetime.datetime, datetime.date)):
                        sub_date = submitted_val.date() if isinstance(submitted_val, datetime.datetime) else submitted_val
                    else:
                        try:
                            sub_date = datetime.datetime.strptime(str(submitted_val).split()[0], "%Y-%m-%d").date()
                        except Exception:
                            pass
                            
                current_job["components"].append({
                    "unit_code": unit_code,
                    "assembly_3d": assembly_3d if assembly_3d else "-",
                    "parts_3d": parts_3d if parts_3d else "-",
                    "assembly_2d": assembly_2d if assembly_2d else "-",
                    "parts_2d": parts_2d if parts_2d else "-",
                    "status": status_val if status_val else "Pending/Not Started",
                    "submitted_date": sub_date
                })
                
        return jobs_dict

    @classmethod
    def _load_and_parse_layout(cls, excel_path: str) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Schedule']
        
        # Get members starting from Row 5 until blank or row 35
        members = []
        for r in range(5, min(ws.max_row + 1, 35)):
            name_val = ws.cell(row=r, column=8).value
            name_str = str(name_val).strip() if name_val is not None else ""
            if name_str and name_str.lower() != "additional members":
                members.append({
                    "row": r,
                    "name": name_str
                })
            elif len(members) >= 5 and not name_str:
                break
                
        # Find the last column that has a day number in row 4
        max_gantt_col = 20
        for c in range(ws.max_column, 19, -1):
            if ws.cell(row=4, column=c).value is not None:
                max_gantt_col = c
                break
                
        raw_days = []
        for c in range(20, max_gantt_col + 1):
            day_num = ws.cell(row=4, column=c).value
            day_week = ws.cell(row=3, column=c).value
            
            # Backtrack to find month
            month_name = ""
            for col_idx in range(c, 19, -1):
                m_val = ws.cell(row=2, column=col_idx).value
                if m_val:
                    month_name = str(m_val).strip()
                    break
                    
            # Cache spreadsheet assignments default values
            default_assignments = {}
            for m in members:
                cell_val = ws.cell(row=m["row"], column=c).value
                default_assignments[m["name"]] = str(cell_val).strip() if cell_val is not None else ""

            # Check cell color in row 4 (day number) to extract default status from template
            day_cell = ws.cell(row=4, column=c)
            default_status = ""
            if day_cell.fill and day_cell.fill.fill_type == "solid" and day_cell.fill.fgColor:
                rgb_val = str(day_cell.fill.fgColor.rgb).upper()
                if "FF0000" in rgb_val:
                    default_status = "Deadline"
                elif "FFC000" in rgb_val:
                    default_status = "Delivered"
                elif "00B0F0" in rgb_val:
                    default_status = "3D"
                elif "FFFF00" in rgb_val:
                    default_status = "2D"
                elif "EB3FC6" in rgb_val:
                    default_status = "Holiday"

            raw_days.append({
                "col_index": c,
                "month": month_name,
                "day": day_num,
                "weekday": day_week,
                "year": determine_year(month_name, day_num, day_week),
                "default_assignments": default_assignments,
                "default_day_status": default_status
            })
            
        # Programmatically extend the timeline to the end of the year for the last parsed year (e.g. 2026)
        if raw_days:
            last_day_obj = raw_days[-1]
            last_year = last_day_obj["year"]
            last_month_name = last_day_obj["month"].lower()
            last_day_num = last_day_obj["day"]
            
            month_map = {
                'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
                'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
            }
            month_names_list = [
                'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December'
            ]
            
            last_m_idx = month_map.get(last_month_name)
            if last_m_idx and (last_m_idx < 12 or last_day_num < 31):
                start_date = datetime.date(last_year, last_m_idx, int(last_day_num)) + datetime.timedelta(days=1)
                end_date = datetime.date(last_year, 12, 31)
                
                curr_date = start_date
                curr_col = last_day_obj["col_index"] + 1
                
                while curr_date <= end_date:
                    m_name = month_names_list[curr_date.month - 1]
                    wd_name = curr_date.strftime("%a")
                    
                    default_assignments = {m["name"]: "" for m in members}
                    
                    raw_days.append({
                        "col_index": curr_col,
                        "month": m_name,
                        "day": curr_date.day,
                        "weekday": wd_name,
                        "year": last_year,
                        "default_assignments": default_assignments,
                        "default_day_status": ""
                    })
                    curr_date += datetime.timedelta(days=1)
                    curr_col += 1

        return {
            "members": [m["name"] for m in members],
            "member_rows": {unicodedata.normalize('NFC', m["name"].strip().lower()): m["row"] for m in members},
            "days": raw_days
        }

    @classmethod
    async def parse_timeline(cls, excel_path: str, db_assignments: List[Any], db_members: List[str] = None) -> Dict[str, Any]:
        # Load and parse layout structure only if not cached
        if cls._cached_layout is None:
            async with cls._lock:
                if cls._cached_layout is None:
                    cls._cached_layout = await asyncio.to_thread(cls._load_and_parse_layout, excel_path)

        # Map database assignments onto the cached layout structure
        db_map = {}
        for a in db_assignments:
            db_map[(a.member_name.strip().lower(), a.col_index)] = a.value
            
        members_to_use = db_members if db_members is not None else cls._cached_layout["members"]
            
        timeline_days = []
        for d in cls._cached_layout["days"]:
            assignments = {}
            for member_name in members_to_use:
                key = (member_name.strip().lower(), d["col_index"])
                if key in db_map:
                    assignments[member_name] = db_map[key] if db_map[key] is not None else ""
                else:
                    assignments[member_name] = d["default_assignments"].get(member_name, "")
                    
            # Explicitly load __day_status__ assignment
            status_key = ('__day_status__', d["col_index"])
            if status_key in db_map:
                assignments['__day_status__'] = db_map[status_key] if db_map[status_key] is not None else ""
            else:
                assignments['__day_status__'] = d.get("default_day_status", "")

            timeline_days.append({
                "col_index": d["col_index"],
                "month": d["month"],
                "day": d["day"],
                "weekday": d["weekday"],
                "year": d["year"],
                "assignments": assignments
            })
            
        return {
            "members": members_to_use,
            "timeline": timeline_days
        }

    @staticmethod
    def generate_excel_export(
        excel_path: str,
        db_components: List[Any],
        db_jobs: List[Any],
        db_assignments: List[Any],
        db_members: List[str],
        target_months: List[str] = None,
        selected_job_ids: List[str] = None
    ) -> BytesIO:
        # Load template from cached bytes (avoids repeated disk reads of the large xlsx)
        if ExcelScheduleService._cached_template_bytes is None:
            with open(excel_path, 'rb') as f:
                ExcelScheduleService._cached_template_bytes = f.read()
        wb = openpyxl.load_workbook(BytesIO(ExcelScheduleService._cached_template_bytes), data_only=False)
        ws = wb['Schedule']

        # Determine the members list (default to template if db_members is None/empty)
        template_members = []
        for r in range(5, 11):
            name_val = ws.cell(row=r, column=8).value
            if name_val:
                template_members.append(str(name_val).strip())
                
        members_to_use = db_members if db_members else template_members
        
        # ── Helper functions defined early so they're available throughout ──────
        def copy_cell_style(src_cell, dst_cell):
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

        # Dynamic row adjustment for members
        def shift_merged_cells_manually(ws, insert_row, amount):
            merged_ranges = list(ws.merged_cells.ranges)
            ws.merged_cells.ranges.clear()
            for r in merged_ranges:
                if r.min_row >= insert_row:
                    r.shift(row_shift=amount, col_shift=0)
                    ws.merged_cells.add(r)
                elif r.max_row >= insert_row:
                    r.max_row += amount
                    ws.merged_cells.add(r)
                else:
                    ws.merged_cells.add(r)

        def shift_row_dimensions_manually(ws, insert_row, amount):
            dims = {}
            for r, dim in list(ws.row_dimensions.items()):
                if r >= insert_row:
                    dims[r + amount] = dim.height
                    del ws.row_dimensions[r]
                else:
                    dims[r] = dim.height
            for r, h in dims.items():
                if h is not None:
                    ws.row_dimensions[r].height = h

        N = len(members_to_use)
        max_member_row = (5 + N) if N > 6 else 10

        _fill_additional_header = openpyxl.styles.PatternFill(fill_type='solid', fgColor='DCE6F1')
        _font_additional_header = openpyxl.styles.Font(name='Arial Unicode MS', size=10, bold=True, color='001F497D')
        _align_left_center = openpyxl.styles.Alignment(horizontal='left', vertical='center')

        ORIGINAL_SET = {
            'loriemar cañete',
            'niñalyn cordova',
            'kerby poniente',
            'zoren ricablanca',
            'jonathan orendain',
            'teoderic limpiado'
        }

        orig_members = []
        add_members = []
        for m in members_to_use:
            norm = unicodedata.normalize('NFC', m.strip().lower())
            if norm in ORIGINAL_SET:
                orig_members.append(m)
            else:
                add_members.append(m)

        if not orig_members and not add_members:
            orig_members = template_members

        num_orig = len(orig_members)
        num_add = len(add_members)

        if num_add > 0:
            insert_at = 5 + max(num_orig, 5)  # Row 10 if 5 original members
            diff = 1 + num_add  # 1 row for "Additional Members" header + num_add member rows
            
            ws.insert_rows(insert_at, diff)
            shift_merged_cells_manually(ws, insert_at, diff)
            shift_row_dimensions_manually(ws, insert_at, diff)

            # Format insert_at as "Additional Members" section header
            ws.row_dimensions[insert_at].height = 22
            for c in range(1, ws.max_column + 1):
                copy_cell_style(ws.cell(row=5, column=c), ws.cell(row=insert_at, column=c))

            _fill_none = openpyxl.styles.PatternFill(fill_type=None)
            for c in range(1, 8):
                ws.cell(row=insert_at, column=c, value=None)
                ws.cell(row=insert_at, column=c).fill = _fill_none

            for c in range(8, 20):
                ws.cell(row=insert_at, column=c).fill = _fill_additional_header

            ws.cell(row=insert_at, column=8, value="Additional Members")
            ws.cell(row=insert_at, column=8).font = _font_additional_header
            ws.cell(row=insert_at, column=8).alignment = _align_left_center
            ws.merge_cells(start_row=insert_at, start_column=8, end_row=insert_at, end_column=19)

            # Copy layout style of row 5 to inserted member rows
            for r in range(insert_at + 1, insert_at + diff):
                ws.row_dimensions[r].height = ws.row_dimensions[5].height
                for c in range(1, ws.max_column + 1):
                    copy_cell_style(ws.cell(row=5, column=c), ws.cell(row=r, column=c))
                ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=19)
                for c in range(1, 8):
                    ws.cell(row=r, column=c, value=None)
                    ws.cell(row=r, column=c).fill = _fill_none

        # Write member names and build member_rows mapping
        member_rows = {}
        for idx, name in enumerate(orig_members):
            r = 5 + idx
            ws.cell(row=r, column=8, value=name)
            norm_name = unicodedata.normalize('NFC', name.strip().lower())
            member_rows[norm_name] = r

        # Clear unused original template rows up to row 9
        for r in range(5 + num_orig, 10):
            ws.cell(row=r, column=8, value=None)

        if num_add > 0:
            insert_at = 5 + max(num_orig, 5)
            for idx, name in enumerate(add_members):
                r = insert_at + 1 + idx
                ws.cell(row=r, column=8, value=name)
                norm_name = unicodedata.normalize('NFC', name.strip().lower())
                member_rows[norm_name] = r

        last_member_row = (10 + num_add) if num_add > 0 else (5 + num_orig - 1)

        _border_top_med = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='000000'),
            right=openpyxl.styles.Side(style='thin', color='000000'),
            top=openpyxl.styles.Side(style='medium', color='000000'),
            bottom=openpyxl.styles.Side(style='thin', color='000000')
        )
        _border_mid_thin = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='000000'),
            right=openpyxl.styles.Side(style='thin', color='000000'),
            top=openpyxl.styles.Side(style='thin', color='000000'),
            bottom=openpyxl.styles.Side(style='thin', color='000000')
        )
        _border_bot_med = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='000000'),
            right=openpyxl.styles.Side(style='thin', color='000000'),
            top=openpyxl.styles.Side(style='thin', color='000000'),
            bottom=openpyxl.styles.Side(style='medium', color='000000')
        )

        for r in range(5, last_member_row + 1):
            if r == 5:
                b_style = _border_top_med
            elif r == last_member_row:
                b_style = _border_bot_med
            else:
                b_style = _border_mid_thin

            for c in range(8, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if type(cell).__name__ != 'MergedCell':
                    cell.border = copy(b_style)

        db_map = {}
        for c in db_components:
            key = (str(c.job_id).strip().lower(), str(c.unit_code).strip().lower())
            db_map[key] = c

        # Status fill colors: Completed=orange/gold, For Checking=green, else no fill
        _FILL_COMPLETED = openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFC000')  # orange/gold
        _FILL_CHECKING  = openpyxl.styles.PatternFill(fill_type='solid', fgColor='92D050')  # green
        _FILL_NONE      = openpyxl.styles.PatternFill(fill_type=None)
        _FONT_BLACK     = openpyxl.styles.Font(name='Arial Unicode MS', color='FF000000', bold=False)

        def apply_status_fill(sheet, row, col, status_text):
            """Apply color fill + black regular font to the status cell and its merged neighbours (cols 12-15)."""
            s = (status_text or '').strip().lower()
            if s in ('completed', 'complete'):
                fill = _FILL_COMPLETED
            elif s in ('for checking', 'checking'):
                fill = _FILL_CHECKING
            else:
                fill = _FILL_NONE
            for c in range(col, col + 4):   # cols 12-15
                cell = sheet.cell(row=row, column=c)
                cell.fill = fill
                cell.font = _FONT_BLACK

        def fmt_pct(val):
            """Convert stored decimal (e.g. 0.9) back to display string (e.g. '90%'), or return '-' for dashes."""
            if val is None:
                return '-'
            s = str(val).strip()
            if s in ('-', '', 'None'):
                return '-'
            try:
                f = float(s)
                if f > 1:
                    return f"{int(round(f))}%"
                return f"{int(round(f * 100))}%"
            except ValueError:
                return s
        # ─────────────────────────────────────────────────────────────────────────
        # Single-pass scan: update existing rows + collect ref rows + find last row
        # (Previously 3 separate O(N) loops over 1200+ rows — now one pass)
        # ─────────────────────────────────────────────────────────────────────────
        current_job_id = None
        seen_jobs      = set()
        ref_header_row = None   # last "Job Status" row seen  → used as template
        ref_comp_row   = None   # a component row with a known status → style ref
        last_used_row  = 1      # last row with any value in cols A-S
        hide_current_job = False
        
        # Precompute selected job ids for fast lookup
        selected_ids_set = None
        if selected_job_ids is not None:
            selected_ids_set = {str(jid).strip().lower() for jid in selected_job_ids}

        for r in range(1, ws.max_row + 1):
            val_a     = ws.cell(row=r, column=1).value
            val_a_str = str(val_a).strip() if val_a is not None else ""

            # Track last non-empty row
            if val_a is not None or any(
                ws.cell(row=r, column=c).value is not None for c in range(2, 20)
            ):
                last_used_row = r

            # Check if this row is a spacer row above a job header by peeking ahead
            # (Optional: to hide spacer rows if the job below it is hidden)
            # For simplicity, we just hide from the header downwards.

            if "Job Status" in val_a_str:
                m = re.match(r"(\d+)\s+Job\s+Status", val_a_str, re.IGNORECASE)
                current_job_id = m.group(1) if m else val_a_str.replace("Job Status", "").strip()
                seen_jobs.add(current_job_id.strip().lower())
                ref_header_row = r   # keep updating → ends up as last header row
                
                if selected_ids_set is not None:
                    if current_job_id.strip().lower() not in selected_ids_set:
                        hide_current_job = True
                        # Also hide the 2 spacer rows above it
                        if r > 2:
                            ws.row_dimensions[r - 1].hidden = True
                            ws.row_dimensions[r - 2].hidden = True
                    else:
                        hide_current_job = False

            if hide_current_job:
                ws.row_dimensions[r].hidden = True

            elif current_job_id:
                if val_a_str in ("", "Machine/Unit Code", "Legend:"):
                    continue

                unit_code = val_a_str
                key = (str(current_job_id).strip().lower(), str(unit_code).strip().lower())

                if key in db_map:
                    db_comp = db_map[key]
                    ws.cell(row=r, column=12, value=db_comp.status)
                    apply_status_fill(ws, r, 12, db_comp.status)
                    if db_comp.submitted_date:
                        ws.cell(row=r, column=16, value=db_comp.submitted_date.strftime("%Y/%m/%d"))
                    else:
                        ws.cell(row=r, column=16, value=None)

                # Track a real component row as style reference
                status_here = ws.cell(row=r, column=12).value
                if ref_comp_row is None and status_here and \
                        str(status_here).strip().lower() in ("pending/not started", "complete", "for checking"):
                    ref_comp_row = r

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        _fill_header = PatternFill(fill_type='solid', fgColor='DCE6F1')
        _fill_sub = PatternFill(fill_type='solid', fgColor='F2F2F2')
        _fill_spacer = PatternFill(fill_type='solid', fgColor='B7DEE8')
        _border_thin = Border(left=Side(style='thin', color='000000'), 
                              right=Side(style='thin', color='000000'), 
                              top=Side(style='thin', color='000000'), 
                              bottom=Side(style='thin', color='000000'))
        _border_bottom_medium = Border(left=Side(style='thin', color='000000'), 
                                       right=Side(style='thin', color='000000'), 
                                       top=Side(style='thin', color='000000'), 
                                       bottom=Side(style='medium', color='000000'))
        _font_bold = Font(name='Arial Unicode MS', bold=True)
        _font_norm = Font(name='Arial Unicode MS', bold=False)
        _align_center = Alignment(horizontal='center', vertical='center')
        
        def _apply_prog_style(cell, fill=None, font=_font_norm, border=_border_thin, align=_align_center):
            if fill:
                cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = align

        last_member_row = (10 + num_add) if num_add > 0 else (5 + num_orig - 1)

        # Remove any merged cell ranges below last_member_row using openpyxl's unmerge_cells
        for m_range in list(ws.merged_cells.ranges):
            if m_range.min_row > last_member_row or m_range.max_row > last_member_row:
                ws.unmerge_cells(m_range.coord)

        # Clear any old cell values, borders, and fills below last_member_row
        _border_none = openpyxl.styles.Border()
        for r in range(last_member_row + 1, max(ws.max_row + 1, last_member_row + 50)):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.border = _border_none
                cell.fill = openpyxl.styles.PatternFill(fill_type=None)

        next_row = last_member_row + 1

        for j in db_jobs:
            j_key = j.job_id.strip().lower()
            if j_key not in seen_jobs:
                # Add 2 light-blue spacer rows (light-blue fill only on cols 1-19, no fill on timeline cols 20+)
                for spacer_offset in range(2):
                    ws.row_dimensions[next_row].height = 20.1
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=next_row, column=c)
                        if type(cell).__name__ != 'MergedCell':
                            cell.value = None
                            cell.fill = _fill_spacer if c <= 19 else _fill_none
                            cell.border = _border_none
                    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=19)
                    next_row += 1
                
                # Header row 1
                for c in range(1, 20):
                    _apply_prog_style(ws.cell(row=next_row, column=c), fill=None, font=_font_bold)
                ws.cell(row=next_row, column=1, value=f"{j.job_id} Job Status")
                
                raw_deadline = str(j.deadline).strip() if j.deadline else ""
                formatted_deadline = re.sub(r"(\d{4})-(\d{2})-(\d{2})", r"\1/\2/\3", raw_deadline)
                clean_deadline = re.sub(r'(?i)^Deadline:\s*', '', formatted_deadline)
                ws.cell(row=next_row, column=2, value=f"Deadline: {clean_deadline}" if clean_deadline else "Deadline:")
                ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=19)
                ws.cell(row=next_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                next_row += 1
                
                # Header row 2
                for c in range(1, 20):
                    _apply_prog_style(ws.cell(row=next_row, column=c), fill=None, font=_font_bold)
                ws.cell(row=next_row, column=1, value="Machine/Unit Code")
                ws.cell(row=next_row, column=2, value="3D")
                ws.cell(row=next_row, column=7, value="2D")
                ws.cell(row=next_row, column=12, value="Status")
                ws.cell(row=next_row, column=16, value="Submitted Date")
                
                ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row + 1, end_column=1)
                ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=6)
                ws.merge_cells(start_row=next_row, start_column=7, end_row=next_row, end_column=11)
                ws.merge_cells(start_row=next_row, start_column=12, end_row=next_row + 1, end_column=15)
                ws.merge_cells(start_row=next_row, start_column=16, end_row=next_row, end_column=19)
                next_row += 1
                
                # Header row 3
                for c in range(1, 20):
                    _apply_prog_style(ws.cell(row=next_row, column=c), fill=None, font=_font_bold)
                ws.cell(row=next_row, column=2, value="Assembly")
                ws.cell(row=next_row, column=5, value="Parts")
                ws.cell(row=next_row, column=7, value="Assembly")
                ws.cell(row=next_row, column=10, value="Parts")
                ws.cell(row=next_row, column=16, value="Date")
                    
                ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=4)
                ws.merge_cells(start_row=next_row, start_column=5, end_row=next_row, end_column=6)
                ws.merge_cells(start_row=next_row, start_column=7, end_row=next_row, end_column=9)
                ws.merge_cells(start_row=next_row, start_column=10, end_row=next_row, end_column=11)
                ws.merge_cells(start_row=next_row, start_column=16, end_row=next_row, end_column=19)
                next_row += 1
                
                j_comps = [c_item for c_item in db_components if str(c_item.job_id).strip().lower() == j_key]
                for c_item in j_comps:
                    is_last = (c_item == j_comps[-1])
                    for c in range(1, 20):
                        border_to_use = _border_bottom_medium if is_last else _border_thin
                        _apply_prog_style(ws.cell(row=next_row, column=c), fill=None, font=_font_norm, border=border_to_use)
                    
                    ws.cell(row=next_row, column=1, value=c_item.unit_code)
                    ws.cell(row=next_row, column=2, value=fmt_pct(c_item.assembly_3d))
                    ws.cell(row=next_row, column=5, value=fmt_pct(c_item.parts_3d))
                    ws.cell(row=next_row, column=7, value=fmt_pct(c_item.assembly_2d))
                    ws.cell(row=next_row, column=10, value=fmt_pct(c_item.parts_2d))
                    ws.cell(row=next_row, column=12, value=c_item.status)
                    apply_status_fill(ws, next_row, 12, c_item.status)
                    if c_item.submitted_date:
                        ws.cell(row=next_row, column=16, value=c_item.submitted_date.strftime("%Y/%m/%d"))
                    else:
                        ws.cell(row=next_row, column=16, value=None)
                        
                    ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=4)
                    ws.merge_cells(start_row=next_row, start_column=5, end_row=next_row, end_column=6)
                    ws.merge_cells(start_row=next_row, start_column=7, end_row=next_row, end_column=9)
                    ws.merge_cells(start_row=next_row, start_column=10, end_row=next_row, end_column=11)
                    ws.merge_cells(start_row=next_row, start_column=12, end_row=next_row, end_column=15)
                    ws.merge_cells(start_row=next_row, start_column=16, end_row=next_row, end_column=19)
                    
                    next_row += 1

        # Find the last column that has a day number in row 4 (Timeline range)
        max_gantt_col = 20
        for c in range(ws.max_column, 19, -1):
            if ws.cell(row=4, column=c).value is not None:
                max_gantt_col = c
                break

        # Dynamically append columns in Excel if assignments exceed max_gantt_col or to complete full 12 months to Dec 31
        ref_day_num = ws.cell(row=4, column=max_gantt_col).value
        ref_day_week = ws.cell(row=3, column=max_gantt_col).value
        ref_month_name = ""
        for col_idx in range(max_gantt_col, 19, -1):
            m_val = ws.cell(row=2, column=col_idx).value
            if m_val:
                ref_month_name = str(m_val).strip()
                break
        ref_year = determine_year(ref_month_name, ref_day_num, ref_day_week)
        month_map = {
            'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
            'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
        }
        ref_m_idx = month_map.get(ref_month_name.lower(), 7)
        try:
            ref_date = datetime.date(ref_year, ref_m_idx, int(ref_day_num))
        except (ValueError, TypeError):
            ref_date = datetime.date(2026, 7, 31)

        dec_31_date = datetime.date(ref_year, 12, 31)
        extra_days_to_year_end = max(0, (dec_31_date - ref_date).days)
        target_year_end_col = max_gantt_col + extra_days_to_year_end

        max_col_to_write = max(max_gantt_col, target_year_end_col, max((a.col_index for a in db_assignments if a.col_index is not None), default=0))

        if max_col_to_write > max_gantt_col:
            for c in range(max_gantt_col + 1, max_col_to_write + 1):
                c_let = openpyxl.utils.get_column_letter(c)
                ref_let = openpyxl.utils.get_column_letter(max_gantt_col)
                ws.column_dimensions[c_let].width = ws.column_dimensions[ref_let].width
                
                days_diff = c - max_gantt_col
                dt = ref_date + datetime.timedelta(days=days_diff)
                
                # Copy styles & write values
                copy_cell_style(ws.cell(row=2, column=max_gantt_col), ws.cell(row=2, column=c))
                if dt.day == 1 or c == max_gantt_col + 1:
                    ws.cell(row=2, column=c, value=dt.strftime("%B"))
                else:
                    ws.cell(row=2, column=c, value=None)
                    
                copy_cell_style(ws.cell(row=3, column=max_gantt_col), ws.cell(row=3, column=c))
                ws.cell(row=3, column=c, value=dt.strftime("%a"))
                
                copy_cell_style(ws.cell(row=4, column=max_gantt_col), ws.cell(row=4, column=c))
                ws.cell(row=4, column=c, value=dt.day)
                
                for r in range(5, max_member_row + 1):
                    copy_cell_style(ws.cell(row=r, column=max_gantt_col), ws.cell(row=r, column=c))

        # Define fills for day status columns (FFFF0000=Deadline, FFFFC000=Delivered, FF00B0F0=3D, FFFFFF00=2D, FFEB3FC6=Holiday)
        _DAY_STATUS_FILLS = {
            'deadline': openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFFF0000'),
            'delivered': openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFFFC000'),
            '3d': openpyxl.styles.PatternFill(fill_type='solid', fgColor='FF00B0F0'),
            '2d': openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFFFFF00'),
            'holiday': openpyxl.styles.PatternFill(fill_type='solid', fgColor='FFEB3FC6'),
        }

        # Pre-clear all timeline colors & values before writing DB state
        # Restores default weekend highlights (theme=6 tint=0.6 for Sat, theme=9 tint=0.6 for Sun) or clears fill
        for c in range(20, max_col_to_write + 1):
            day_val = str(ws.cell(row=3, column=c).value or '').strip().lower()
            if 'sat' in day_val:
                default_fill = openpyxl.styles.PatternFill(
                    fill_type='solid',
                    fgColor=openpyxl.styles.colors.Color(theme=6, tint=0.5999938962981048)
                )
            elif 'sun' in day_val:
                default_fill = openpyxl.styles.PatternFill(
                    fill_type='solid',
                    fgColor=openpyxl.styles.colors.Color(theme=9, tint=0.5999938962981048)
                )
            else:
                default_fill = openpyxl.styles.PatternFill(fill_type=None)
            
            for r in range(3, max_member_row + 1):
                ws.cell(row=r, column=c).fill = default_fill

            for r in range(5, max_member_row + 1):
                ws.cell(row=r, column=c, value=None)

        # Write database assignments and column statuses
        # 1. Day statuses first
        for a in db_assignments:
            m_name_lower = a.member_name.strip().lower()
            if m_name_lower == '__day_status__':
                status_key = (a.value or '').strip().lower()
                if status_key in _DAY_STATUS_FILLS:
                    fill = _DAY_STATUS_FILLS[status_key]
                    for r in range(3, max_member_row + 1):
                        ws.cell(row=r, column=a.col_index).fill = fill

        # 2. Member assignments grouped by member
        member_assignments = {}
        for a in db_assignments:
            m_name_lower = a.member_name.strip().lower()
            if m_name_lower == '__day_status__' or not a.value:
                continue
            member_key = unicodedata.normalize('NFC', a.member_name.strip().lower())
            if member_key in member_rows:
                if member_key not in member_assignments:
                    member_assignments[member_key] = []
                member_assignments[member_key].append(a)

        # Style objects for arrow components and text
        _FONT_RED = openpyxl.styles.Font(name='Arial Unicode MS', size=9, color='FFFF0000', bold=True)
        _FONT_BLUE = openpyxl.styles.Font(name='Arial Unicode MS', size=9, color='FF0070C0', bold=True)
        _ALIGN_CENTER = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Keep timeline cells clean and unmerged so DrawingML floating shapes (arrows and text boxes) display cleanly
        for member_key, ass_list in member_assignments.items():
            target_row = member_rows[member_key]
            for a in ass_list:
                ws.cell(row=target_row, column=a.col_index, value=None)
            
        # ─────────────────────────────────────────────────────────────────────────
        # ─────────────────────────────────────────────────────────────────────────
        # 12-Month Bounds & Target Month Filtering
        # ─────────────────────────────────────────────────────────────────────────
        # ─────────────────────────────────────────────────────────────────────────
        # 12-Month Bounds & Target Month Filtering
        # ─────────────────────────────────────────────────────────────────────────
        # Build mapping of column -> year
        col_years = {}
        for c in range(20, max_col_to_write + 1):
            day_num = ws.cell(row=4, column=c).value
            day_week = ws.cell(row=3, column=c).value
            month_name = None
            for col_idx in range(c, 19, -1):
                m_val = ws.cell(row=2, column=col_idx).value
                if m_val:
                    month_name = str(m_val).strip()
                    break
            if day_num and day_week and month_name:
                col_years[c] = determine_year(month_name, day_num, day_week)

        # Detect active year from db_assignments or default to current calendar year
        active_col_indices = [a.col_index for a in db_assignments if a.col_index and a.col_index >= 20 and a.value and a.member_name != '__day_status__']
        target_year = datetime.date.today().year

        if active_col_indices:
            assigned_years = [col_years.get(c) for c in active_col_indices if c in col_years]
            assigned_years = [y for y in assigned_years if y is not None]
            if assigned_years:
                target_year = max(set(assigned_years), key=assigned_years.count)

        target_year_cols = [c for c, y in col_years.items() if y == target_year]
        if target_year_cols:
            start_year_col = min(target_year_cols)
            end_year_col = max(target_year_cols)
        else:
            start_year_col = 20
            end_year_col = min(384, max_col_to_write)

        # Calculate dynamic column widths for short duration spans (e.g. 1 or 2 days) with long job labels
        col_widths = {}
        for member_key, ass_list in member_assignments.items():
            ass_list.sort(key=lambda x: x.col_index)
            blocks = []
            if ass_list:
                curr = [ass_list[0]]
                for next_a in ass_list[1:]:
                    if next_a.col_index - curr[-1].col_index <= 1:
                        curr.append(next_a)
                    else:
                        blocks.append(curr)
                        curr = [next_a]
                blocks.append(curr)

            for block in blocks:
                job_a = next((x for x in block if x.value not in ('-->', '->', '---')), None)
                if not job_a:
                    continue
                label = str(job_a.value).strip()
                req_width = len(label) * 0.95 + 1.5
                span_len = len(block)
                per_col_req = req_width / span_len
                for a in block:
                    col_widths[a.col_index] = max(col_widths.get(a.col_index, 3.50), per_col_req)

        # Unhide columns for the target year (e.g. 2026), and set width dynamically
        for c in range(20, max_col_to_write + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            if start_year_col <= c <= end_year_col:
                ws.column_dimensions[col_letter].hidden = False
                ws.column_dimensions[col_letter].outline_level = 0
                target_w = col_widths.get(c, 3.50)
                ws.column_dimensions[col_letter].width = max(3.50, round(target_w, 2))
            else:
                ws.column_dimensions[col_letter].hidden = True

        # Filter by requested target months within the target year if specified
        if target_months:
            target_months_lower = {m.strip().lower() for m in target_months}
            for c in range(start_year_col, end_year_col + 1):
                m_val = None
                for col_idx in range(c, 19, -1):
                    val = ws.cell(row=2, column=col_idx).value
                    if val:
                        m_val = str(val).strip().lower()
                        break
                if m_val and m_val not in target_months_lower:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    ws.column_dimensions[col_letter].hidden = True

        # Set scroll position for unfrozen pane (topLeftCell)
        target_scroll_col = None

        # 1. If target_months specified, scroll to first target month in the target year
        if target_months:
            first_m = target_months[0].strip().lower()
            for c in range(start_year_col, end_year_col + 1):
                m_val = None
                for col_idx in range(c, 19, -1):
                    val = ws.cell(row=2, column=col_idx).value
                    if val:
                        m_val = str(val).strip().lower()
                        break
                if m_val == first_m:
                    target_scroll_col = c
                    break

        # 2. Default scroll position for 12-month export is the CURRENT calendar month (e.g. July) in target year
        if not target_scroll_col:
            current_month_name = datetime.date.today().strftime("%B").lower()
            for c in range(start_year_col, end_year_col + 1):
                m_val = None
                for col_idx in range(c, 19, -1):
                    val = ws.cell(row=2, column=col_idx).value
                    if val:
                        m_val = str(val).strip().lower()
                        break
                if m_val == current_month_name:
                    target_scroll_col = c
                    break

        # 3. Fallback to start of target year if current month not found
        if not target_scroll_col:
            target_scroll_col = start_year_col

        if ws.sheet_view and ws.sheet_view.pane:
            target_col_letter = openpyxl.utils.get_column_letter(target_scroll_col)
            ws.sheet_view.pane.topLeftCell = f"{target_col_letter}1"

        # ─────────────────────────────────────────────────────────────────────────
        # Complete Excel Export with DrawingML Floating Shapes (Arrows & TextBoxes)
        # ─────────────────────────────────────────────────────────────────────────
        out_base = BytesIO()
        wb.save(out_base)
        out_base.seek(0)

        # Load template drawing1.xml and drawing1.xml.rels
        template_drawing_bytes = None
        template_rels_bytes = None
        if os.path.exists(excel_path):
            try:
                with zipfile.ZipFile(excel_path, 'r') as ztmpl:
                    if 'xl/drawings/drawing1.xml' in ztmpl.namelist():
                        template_drawing_bytes = ztmpl.read('xl/drawings/drawing1.xml')
                    if 'xl/drawings/_rels/drawing1.xml.rels' in ztmpl.namelist():
                        template_rels_bytes = ztmpl.read('xl/drawings/_rels/drawing1.xml.rels')
            except Exception:
                pass

        if not template_drawing_bytes:
            return out_base

        try:
            updated_drawing_bytes = update_drawing_xml(template_drawing_bytes, db_assignments, member_rows, col_widths)
            out_final = BytesIO()
            with zipfile.ZipFile(out_base, 'r') as zin:
                with zipfile.ZipFile(out_final, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                    has_drawing = False
                    has_drawing_rels = False
                    has_sheet_rels = False

                    sheet_rels_content = (
                        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
                        '</Relationships>'
                    )

                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/worksheets/sheet1.xml':
                            xml_str = data.decode('utf-8')
                            if 'xmlns:r=' not in xml_str:
                                xml_str = xml_str.replace('<worksheet', '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"', 1)
                            if '<drawing' not in xml_str:
                                xml_str = xml_str.replace('</worksheet>', '<drawing r:id="rId2"/></worksheet>')
                            zout.writestr(item.filename, xml_str.encode('utf-8'))

                        elif item.filename == 'xl/worksheets/_rels/sheet1.xml.rels':
                            xml_str = data.decode('utf-8')
                            if 'drawing1.xml' not in xml_str:
                                xml_str = xml_str.replace(
                                    '</Relationships>',
                                    '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>'
                                )
                            zout.writestr(item.filename, xml_str.encode('utf-8'))
                            has_sheet_rels = True

                        elif item.filename == '[Content_Types].xml':
                            xml_str = data.decode('utf-8')
                            if 'drawing1.xml' not in xml_str:
                                xml_str = xml_str.replace(
                                    '</Types>',
                                    '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>'
                                )
                            zout.writestr(item.filename, xml_str.encode('utf-8'))

                        elif item.filename == 'xl/drawings/drawing1.xml':
                            zout.writestr(item.filename, updated_drawing_bytes)
                            has_drawing = True

                        elif item.filename == 'xl/drawings/_rels/drawing1.xml.rels':
                            if template_rels_bytes:
                                zout.writestr(item.filename, template_rels_bytes)
                                has_drawing_rels = True
                            else:
                                zout.writestr(item.filename, data)
                        else:
                            zout.writestr(item.filename, data)

                    if not has_drawing and updated_drawing_bytes:
                        zout.writestr('xl/drawings/drawing1.xml', updated_drawing_bytes)
                    if not has_drawing_rels and template_rels_bytes:
                        zout.writestr('xl/drawings/_rels/drawing1.xml.rels', template_rels_bytes)
                    if not has_sheet_rels:
                        zout.writestr('xl/worksheets/_rels/sheet1.xml.rels', sheet_rels_content)

            out_final.seek(0)
            return out_final
        except Exception as err:
            print(f"[EXCEL-EXPORT-WARNING] Failed to inject DrawingML shapes: {err}")
            out_base.seek(0)
            return out_base
